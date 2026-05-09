from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models import DeviationResult, RiskItem


NEGATIVE_FILL = PatternFill("solid", fgColor="FCE4E4")
POSITIVE_FILL = PatternFill("solid", fgColor="E6F4EA")
LOW_CONFIDENCE_FILL = PatternFill("solid", fgColor="FFF4CE")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def export_deviation_table(
    path: str | Path,
    deviations: list[DeviationResult],
    risks: list[RiskItem] | None = None,
    include_confidence: bool = False,
) -> Path:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "技术偏离表"

    headers = ["序号", "招标要求", "我方响应", "偏离情况", "证明材料", "来源页码"]
    if include_confidence:
        headers.append("置信度")
    sheet.append(headers)
    _style_header(sheet)

    for index, item in enumerate(deviations, start=1):
        row = [
            index,
            f"{item.item} - {item.parameter}: {item.required_value}",
            item.response_text,
            _deviation_label(item.deviation_type.value),
            item.evidence or "",
            item.source_page or "",
        ]
        if include_confidence:
            row.append(item.confidence)
        sheet.append(row)
        _style_deviation_row(sheet, sheet.max_row, item)

    _fit_columns(sheet)

    if risks is not None:
        risk_sheet = workbook.create_sheet("废标风险")
        risk_sheet.append(["序号", "等级", "类型", "要求", "建议", "来源页码", "状态"])
        _style_header(risk_sheet)
        for index, risk in enumerate(risks, start=1):
            risk_sheet.append(
                [
                    index,
                    risk.severity.value,
                    risk.risk_type,
                    risk.requirement,
                    risk.suggestion,
                    risk.source_page or "",
                    risk.status.value,
                ]
            )
        _fit_columns(risk_sheet)

    workbook.save(target)
    return target


def export_scoring_matrix(path: str | Path, rows: list[dict]) -> Path:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "评分点响应矩阵"
    headers = ["序号", "来源", "评分点", "要求", "响应", "证明材料", "来源页码", "状态", "优先级"]
    sheet.append(headers)
    _style_header(sheet)
    for index, row in enumerate(rows, start=1):
        sheet.append(
            [
                index,
                row.get("source", ""),
                row.get("score_point", ""),
                row.get("requirement", ""),
                row.get("response", ""),
                row.get("evidence", ""),
                row.get("source_page", ""),
                row.get("status", ""),
                row.get("priority", ""),
            ]
        )
    _fit_columns(sheet)
    workbook.save(target)
    return target


def export_material_gap_list(path: str | Path, rows: list[dict]) -> Path:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "材料缺失清单"
    headers = ["序号", "等级", "风险类型", "材料要求", "建议", "已绑定材料", "来源页码", "来源章节", "状态"]
    sheet.append(headers)
    _style_header(sheet)
    for index, row in enumerate(rows, start=1):
        bound_materials = row.get("bound_materials") or row.get("material_ids") or []
        sheet.append(
            [
                index,
                row.get("severity", ""),
                row.get("risk_type", ""),
                row.get("material_requirement", ""),
                row.get("suggestion", ""),
                ", ".join(str(item) for item in bound_materials),
                row.get("source_page", ""),
                row.get("source_section", ""),
                row.get("status", ""),
            ]
        )
    _fit_columns(sheet)
    workbook.save(target)
    return target


def _style_header(sheet) -> None:
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def _style_deviation_row(sheet, row_number: int, item: DeviationResult) -> None:
    fill = None
    if item.confidence < 0.7:
        fill = LOW_CONFIDENCE_FILL
    elif item.deviation_type.value == "negative":
        fill = NEGATIVE_FILL
    elif item.deviation_type.value == "positive":
        fill = POSITIVE_FILL
    if fill is None:
        return
    for cell in sheet[row_number]:
        cell.fill = fill


def _fit_columns(sheet) -> None:
    for column in sheet.columns:
        width = 10
        for cell in column:
            width = max(width, min(len(str(cell.value or "")) + 2, 48))
        sheet.column_dimensions[get_column_letter(column[0].column)].width = width


def _deviation_label(value: str) -> str:
    return {
        "positive": "正偏离",
        "none": "无偏离",
        "negative": "负偏离",
        "unknown": "待判断",
    }.get(value, value)
