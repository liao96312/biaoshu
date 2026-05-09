import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from app.models import Material, RiskItem, Severity, TechRequirement
from app.services.checklists import build_material_gap_list, build_scoring_matrix
from app.services.clauses import extract_clauses
from app.services.deviation import build_deviation, judge_deviation
from app.services.exporter import export_deviation_table, export_material_gap_list, export_scoring_matrix
from app.services.pdf_report import export_risk_report_pdf
from app.services.product_matcher import match_products
from app.services.risk_report import export_risk_report_docx
from app.services.risk_scanner import scan_risks
from app.services.tech_params import extract_tech_requirements


class ServiceTests(unittest.TestCase):
    def test_scan_risks_keeps_source_page(self):
        text = "[[page:12]]\n投标人须提供投标保证金缴纳凭证，否则按无效投标处理。"
        risks = scan_risks(text, "proj_1")
        self.assertEqual(len(risks), 1)
        self.assertEqual(risks[0].source_page, 12)
        self.assertEqual(risks[0].severity.value, "high")

    def test_extract_tech_requirement(self):
        text = "[[page:23]]\n★核心交换机整机交换容量≥2.56Tbps，包转发率≥720Mpps。"
        requirements = extract_tech_requirements(text, "proj_1")
        self.assertGreaterEqual(len(requirements), 2)
        self.assertEqual(requirements[0].operator, ">=")
        self.assertTrue(requirements[0].is_mandatory)

    def test_extract_clauses_keeps_page_section_and_type(self):
        text = "[[page:5]]\n第四章 技术要求\n投标人须提供投标保证金缴纳凭证，否则投标无效。\n★核心交换机包转发率≥720Mpps。"
        clauses = extract_clauses(text)
        self.assertGreaterEqual(len(clauses), 2)
        self.assertEqual(clauses[0]["source_page"], 5)
        self.assertEqual(clauses[0]["source_section"], "第四章 技术要求")
        self.assertEqual(clauses[0]["clause_type"], "risk")
        self.assertEqual(clauses[1]["clause_type"], "technical")

    def test_judge_deviation_by_program(self):
        self.assertEqual(judge_deviation(">=", 720, "720Mpps").value, "none")
        self.assertEqual(judge_deviation(">=", 720, "800Mpps").value, "positive")
        self.assertEqual(judge_deviation(">=", 720, "600Mpps").value, "negative")
        self.assertEqual(judge_deviation(">=", 2.56, "3200Gbps", "Tbps").value, "positive")

    def test_export_xlsx(self):
        requirement = TechRequirement(
            id="P-001",
            project_id="proj_1",
            item_name="核心交换机",
            parameter_name="包转发率",
            operator=">=",
            required_value=720,
            unit="Mpps",
            source_page=23,
            source_text="包转发率≥720Mpps",
        )
        deviation = build_deviation(requirement, "800Mpps")
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "deviation.xlsx"
            export_deviation_table(path, [deviation])
            workbook = load_workbook(path)
            self.assertIn("技术偏离表", workbook.sheetnames)
            self.assertEqual(workbook["技术偏离表"]["D2"].value, "正偏离")

    def test_product_matcher_fills_our_value(self):
        requirement = extract_tech_requirements(
            "[[page:1]]\n核心交换机包转发率≥720Mpps。", "proj_1"
        )[0]
        material = Material(
            id="mat_1",
            company_id="comp_1",
            file_name="产品规格书.pdf",
            material_type="product",
            storage_path="",
            parsed_text="[[page:3]]\n核心交换机包转发率≥800Mpps。",
        )
        result = match_products([requirement], [material])[0]
        self.assertEqual(result.our_value, "800Mpps")
        self.assertEqual(result.deviation_type.value, "positive")

    def test_product_matcher_allows_convertible_units(self):
        from app.models import Material

        requirement = extract_tech_requirements(
            "[[page:1]]\n核心交换机整机交换容量≥2.56Tbps。", "proj_1"
        )[0]
        material = Material(
            id="mat_1",
            company_id="comp_1",
            file_name="产品规格书.pdf",
            material_type="product",
            storage_path="",
            parsed_text="[[page:3]]\n核心交换机整机交换容量≥3200Gbps。",
        )
        result = match_products([requirement], [material])[0]
        self.assertEqual(result.deviation_type.value, "positive")

    def test_export_risk_report_docx(self):
        from zipfile import ZipFile
        from app.models import Project

        project = Project(id="proj_1", name="项目", tender_name="招标", company_id="comp_1")
        risks = scan_risks("[[page:2]]\n投标人须盖章，否则投标无效。", project.id)
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "risk.docx"
            export_risk_report_docx(path, project, risks)
            with ZipFile(path) as docx:
                self.assertIn("word/document.xml", docx.namelist())
            pdf_path = Path(temp_dir) / "risk.pdf"
            export_risk_report_pdf(pdf_path, project, risks)
            self.assertTrue(pdf_path.read_bytes().startswith(b"%PDF-1.4"))

    def test_material_gaps_and_scoring_matrix(self):
        risks = scan_risks("[[page:2]]\n投标人须提供投标保证金缴纳凭证，否则投标无效。", "proj_1")
        gaps = build_material_gap_list(risks)
        self.assertGreaterEqual(len(gaps), 1)

        requirement = TechRequirement(
            id="P-001",
            project_id="proj_1",
            item_name="核心交换机",
            parameter_name="包转发率",
            operator=">=",
            required_value=720,
            unit="Mpps",
            source_page=23,
            source_text="包转发率≥720Mpps",
        )
        deviation = build_deviation(requirement, "600Mpps")
        matrix = build_scoring_matrix([deviation], risks)
        self.assertEqual(matrix[0]["priority"], "high")
        bound_risk = RiskItem(
            id="risk_bound",
            project_id="proj_1",
            risk_type="符合性废标",
            requirement="须提供投标保证金缴纳凭证",
            trigger_keyword="须提供",
            severity=Severity.HIGH,
            source_text="须提供投标保证金缴纳凭证，否则投标无效",
            ai_reason="命中规则",
            suggestion="上传凭证",
            confidence=0.9,
            material_ids=["mat_1"],
        )
        material = Material(
            id="mat_1",
            company_id="comp_1",
            file_name="guarantee.txt",
            material_type="qualification",
            storage_path="",
            name="投标保证金回单",
        )
        bound_gaps = build_material_gap_list([bound_risk], [material])
        self.assertEqual(bound_gaps[0]["bound_materials"], ["投标保证金回单"])
        self.assertIn("投标保证金回单", build_scoring_matrix([], [bound_risk], [material])[0]["evidence"])
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "scoring.xlsx"
            export_scoring_matrix(path, matrix)
            workbook = load_workbook(path)
            self.assertIn("评分点响应矩阵", workbook.sheetnames)
            gap_path = Path(temp_dir) / "gaps.xlsx"
            export_material_gap_list(gap_path, bound_gaps)
            workbook = load_workbook(gap_path)
            self.assertIn("材料缺失清单", workbook.sheetnames)
            self.assertEqual(workbook["材料缺失清单"]["F2"].value, "投标保证金回单")


if __name__ == "__main__":
    unittest.main()
